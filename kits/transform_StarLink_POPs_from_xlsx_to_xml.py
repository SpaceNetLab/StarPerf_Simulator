'''

Author : yunanhou

Date : 2023/11/14

Function : convert Starlink POP data information from xlsx format to xml format

'''

import openpyxl
from xml.etree.ElementTree import Element, SubElement, tostring
from xml.dom import minidom


def xlsx_to_xml(xlsx_file_path, xml_file_path):
    # read the Excel file
    workbook = openpyxl.load_workbook(xlsx_file_path)
    sheet = workbook.active
    # create XML root element
    root = Element('POPs')
    # get header field name
    header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    # traverse the data row starting with the second row
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
        gs_element = SubElement(root, f'POP{row_index}')
        for index, value in enumerate(row):
            field_name = header_row[index]
            field = SubElement(gs_element, field_name)
            field.text = str(value)
    # format using minidom
    xml_string = minidom.parseString(tostring(root)).toprettyxml(indent="  ")
    # delete the first line
    xml_lines = xml_string.split('\n')[1:]
    formatted_xml = '\n'.join(xml_lines)
    # write to XML file
    with open(xml_file_path, 'w' , encoding="utf-8") as xml_file:
        xml_file.write(formatted_xml)
